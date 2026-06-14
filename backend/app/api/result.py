import io, os, tempfile
from fastapi import APIRouter, Depends, Query, UploadFile, File, HTTPException, Request
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.database import get_db
from app.middleware.auth import get_current_user
from app.schemas.result import ResultCreate, ResultOut, ResultQueryRequest, ResultFilter
from app.services import result_service
from app.utils.audit import log_event

admin_router = APIRouter(prefix="/api/admin/results", tags=["成绩管理"])
public_router = APIRouter(prefix="/api/public/contests", tags=["前台成绩"])


@admin_router.get("", response_model=dict)
async def list_results(
    contest_id: int | None = Query(default=None),
    group_id: int | None = Query(default=None),
    is_published: bool | None = Query(default=None),
    keyword: str = Query(default=""),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    filters = ResultFilter(
        contest_id=contest_id, group_id=group_id, is_published=is_published,
        keyword=keyword, page=page, page_size=page_size,
    )
    items, total = await result_service.list_results(db, filters)
    return {
        "items": [
            {
                **ResultOut.model_validate(r).model_dump(),
                "registration_number": r.registration.registration_number if r.registration else "",
                "contestant_name": r.registration.form_data.get("name", "") if r.registration else "",
            }
            for r in items
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@admin_router.get("/template")
async def download_template(
    contest_id: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from app.models.contest import Contest
    from app.models.registration import Registration
    from sqlalchemy import select as sa_select

    r = await db.execute(sa_select(Contest).where(Contest.id == contest_id))
    contest = r.scalar_one_or_none()
    cats = contest.score_categories if contest and contest.score_categories else ["客观题得分", "主观题得分"]

    # Get registrations
    regs_r = await db.execute(
        sa_select(Registration).where(Registration.contest_id == contest_id, Registration.deleted_at.is_(None))
        .order_by(Registration.submitted_at)
    )
    registrations = list(regs_r.scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = "成绩导入模板"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    clean_cats = [c for c in cats if c.strip()]
    headers = ["报名编号", "姓名"] + clean_cats + ["总分", "排名", "奖项"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # Pre-fill registration numbers and names
    for row_idx, reg in enumerate(registrations, 2):
        ws.cell(row=row_idx, column=1, value=reg.registration_number)
        ws.cell(row=row_idx, column=2, value=reg.form_data.get("name", ""))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    from urllib.parse import quote
    from datetime import datetime, timezone
    safe_title = contest.title.replace('/', '_').replace('\\', '_') if contest else f"赛事{contest_id}"
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"{safe_title}_成绩导入模板_{timestamp}.xlsx"

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    return FileResponse(tmp.name, filename=filename,
                       media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Import helpers ───────────────────────────────────────────────


def _parse_import_headers(ws: Worksheet) -> tuple[list[tuple[int, str]], int, int, int]:
    """Parse the header row of an import worksheet.

    Returns:
        (score_cols, total_col, rank_col, award_col) where score_cols is [(col_index, col_name), ...].
    """
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    non_score = {"报名编号", "姓名", "总分", "排名", "奖项"}
    score_cols = []
    total_col = -1
    rank_col = -1
    award_col = -1
    for i, h in enumerate(headers):
        if h == "总分":
            total_col = i
        elif h == "排名":
            rank_col = i
        elif h == "奖项":
            award_col = i
        elif h not in non_score:
            score_cols.append((i, h))
    return score_cols, total_col, rank_col, award_col


async def _process_import_row(
    db: AsyncSession, row: tuple, row_idx: int, contest_id: int,
    score_cols: list[tuple[int, str]], total_col: int, rank_col: int, award_col: int,
) -> tuple[bool, str]:
    """Process a single import row. Returns (success, error_message)."""
    from sqlalchemy import select
    from app.models.registration import Registration
    from app.models.contest import Award

    reg_number = str(row[0]).strip()

    r = await db.execute(select(Registration).where(
        Registration.registration_number == reg_number,
        Registration.deleted_at.is_(None),
    ))
    reg = r.scalar_one_or_none()
    if not reg:
        return False, f"报名编号 {reg_number} 不存在"

    # Build scores dict from dynamic columns
    scores = {}
    for col_idx, col_name in score_cols:
        if len(row) > col_idx and row[col_idx] is not None:
            try:
                scores[col_name] = float(row[col_idx])
            except (ValueError, TypeError):
                return False, f"第{col_idx + 1}列「{col_name}」不是有效数字"

    # Total score: use provided value or sum of score columns
    try:
        total = float(row[total_col]) if total_col >= 0 and len(row) > total_col and row[total_col] is not None else sum(scores.values())
    except (ValueError, TypeError):
        return False, "总分不是有效数字"

    # Rank
    rank = None
    if rank_col >= 0 and len(row) > rank_col and row[rank_col] is not None:
        try:
            rank = int(row[rank_col])
        except (ValueError, TypeError):
            return False, "排名不是有效整数"

    # Award
    award_id = None
    if award_col >= 0 and len(row) > award_col and row[award_col] is not None:
        award_name = str(row[award_col]).strip()
        if award_name:
            a = await db.execute(select(Award).where(Award.contest_id == contest_id, Award.name == award_name))
            award = a.scalar_one_or_none()
            if award:
                award_id = award.id
            else:
                return False, f"奖项「{award_name}」不存在"

    data = ResultCreate(
        contest_id=contest_id,
        registration_id=reg.id,
        scores=scores,
        total_score=total,
        rank=rank,
        award_id=award_id,
    )
    await result_service.create_or_update_result(db, data)
    return True, ""


# ── Import endpoint ──────────────────────────────────────────────


@admin_router.post("/import")
async def import_results(
    request: Request,
    contest_id: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Import results from an .xlsx file. Expects header row with 报名编号, 姓名, score columns, 总分, 排名, 奖项."""
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 格式文件")

    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active

    score_cols, total_col, rank_col, award_col = _parse_import_headers(ws)

    success_count = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row[0]:
            continue  # Skip empty rows

        ok, error_msg = await _process_import_row(
            db, row, row_idx, contest_id, score_cols, total_col, rank_col, award_col,
        )
        if ok:
            success_count += 1
        else:
            errors.append({"row": row_idx, "error": error_msg})

    await log_event(db, "import_results", operator=current_user["username"], operator_id=current_user["user_id"],
                    target=f"contest/{contest_id}", target_type="result",
                    detail={"success_count": success_count, "error_count": len(errors)},
                    result="success", request=request)
    return {"success_count": success_count, "error_count": len(errors), "errors": errors[:20]}


@admin_router.post("", response_model=ResultOut)
async def create_result(data: ResultCreate, request: Request, db: AsyncSession = Depends(get_db), current_user: dict = Depends(get_current_user)):
    result = await result_service.create_or_update_result(db, data)
    await log_event(db, "upsert_result", operator=current_user["username"], operator_id=current_user["user_id"],
                    target=f"contest/{data.contest_id}/reg/{data.registration_id}", target_type="result",
                    result="success", request=request)
    return result


@admin_router.patch("/{result_id}/publish")
async def publish_result(result_id: int, request: Request, db: AsyncSession = Depends(get_db), current_user: dict = Depends(get_current_user)):
    await result_service.publish_result(db, result_id)
    await log_event(db, "publish_result", operator=current_user["username"], operator_id=current_user["user_id"],
                    target=str(result_id), target_type="result", result="success", request=request)
    return {"message": "成绩已发布"}


@admin_router.patch("/{result_id}/withdraw")
async def withdraw_result(result_id: int, request: Request, db: AsyncSession = Depends(get_db), current_user: dict = Depends(get_current_user)):
    await result_service.withdraw_result(db, result_id)
    await log_event(db, "withdraw_result", operator=current_user["username"], operator_id=current_user["user_id"],
                    target=str(result_id), target_type="result", result="success", request=request)
    return {"message": "成绩已撤回"}


# --- Public ---

@public_router.post("/{contest_id}/query-result")
async def query_result(contest_id: int, data: ResultQueryRequest, db: AsyncSession = Depends(get_db)):
    result = await result_service.query_result_public(db, contest_id, data.registration_number, data.email)
    if not result:
        raise HTTPException(status_code=404, detail="未查询到成绩，请检查报名编号和手机号是否正确")
    return result
