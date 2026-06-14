import os
import uuid
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.services.registration_service import get_registrations_for_export
from app.services.result_service import list_results
from app.schemas.result import ResultFilter
from app.utils.crypto import decrypt_value

# Chinese field name mapping
FIELD_LABELS = {
    "registration_number": "报名编号",
    "name": "姓名",
    "email": "邮箱",
    "id_number": "身份证号",
    "organization": "学校/单位",
    "group_id": "组别",
    "group": "组别",
    "submitted_at": "报名时间",
    "total_score": "总分",
    "rank": "排名",
    "award": "奖项",
    "scores": "评分详情",
}

# In-memory export task store
_export_tasks: dict[str, dict] = {}

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
body_font = Font(name="微软雅黑", size=10)
center_align = Alignment(horizontal="center", vertical="center")


# ── Private helpers ──────────────────────────────────────────────


def _resolve_default_field_names(export_type: str, score_categories: list[str] | None = None) -> list[str]:
    """Return the default field name list for a given export type.

    For registration exports: fixed fields (registration_number, name, email, etc.).
    For result exports: base fields plus dynamic score category columns from the contest.
    """
    if export_type == "registration":
        return ["registration_number", "name", "email", "id_number", "organization", "group_id", "submitted_at"]

    cats = [c for c in (score_categories or []) if c.strip()]
    return ["registration_number", "name"] + cats + ["total_score", "rank", "award"]


def _write_header_row(ws: Worksheet, field_names: list[str]) -> None:
    """Write the styled header row with Chinese labels for the given field names."""
    for col, name in enumerate(field_names, 1):
        label = FIELD_LABELS.get(name, name)
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def _write_registration_rows(
    ws: Worksheet, rows: list, field_names: list[str]
) -> None:
    """Fill registration data rows into the worksheet starting at row 2."""
    for row_idx, reg in enumerate(rows, 2):
        form_data = reg.form_data or {}
        for col_idx, name in enumerate(field_names, 1):
            if name == "registration_number":
                val = reg.registration_number
            elif name == "name":
                val = form_data.get("name", "")
            elif name == "email":
                val = form_data.get("email", "")
            elif name == "id_number":
                val = decrypt_value(form_data.get("id_number", ""))
            elif name == "organization":
                val = form_data.get("organization", "")
            elif name == "group_id":
                val = str(reg.group_id) if reg.group_id else ""
            elif name == "submitted_at":
                val = reg.submitted_at.strftime("%Y-%m-%d %H:%M") if reg.submitted_at else ""
            else:
                val = form_data.get(name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
            cell.font = body_font
            cell.border = thin_border


async def _write_result_rows(
    ws: Worksheet, items: list, db: AsyncSession, field_names: list[str]
) -> None:
    """Fill result data rows, looking up per-row registration info for name/email fields."""
    from app.services.registration_service import get_registration

    for row_idx, r in enumerate(items, 2):
        # Best-effort registration lookup: missing registration shouldn't block the whole export
        try:
            reg = await get_registration(db, r.registration_id)
            form_data = reg.form_data or {}
        except Exception:
            form_data = {}

        for col_idx, name in enumerate(field_names, 1):
            if name == "registration_number":
                val = getattr(r, 'registration_number', '') if hasattr(r, 'registration_number') else ''
                if not val and hasattr(r, 'registration'):
                    val = r.registration.registration_number if r.registration else ''
            elif name == "name":
                val = form_data.get("name", "")
            elif name == "email":
                val = form_data.get("email", "")
            elif name == "id_number":
                val = decrypt_value(form_data.get("id_number", ""))
            elif name == "organization":
                val = form_data.get("organization", "")
            elif name == "total_score":
                val = str(r.total_score)
            elif name == "rank":
                val = str(r.rank) if r.rank else ""
            elif name == "award":
                val = ""
            else:
                val = str(r.scores.get(name, "")) if r.scores else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
            cell.font = body_font
            cell.border = thin_border


def _auto_fit_columns(ws: Worksheet) -> None:
    """Set column widths to fit content, capped at 50 characters.

    Best-effort: silently skips cells whose value cannot be stringified.
    """
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


# ── Public API ───────────────────────────────────────────────────


async def submit_export_task(
    db: AsyncSession, export_type: str, contest_id: int, field_names: list[str],
    group_ids: list[int] | None = None,
) -> str:
    """Create an async export task: fetch data, render Excel, and save to disk.

    Returns a task_id for polling via get_export_task_status.
    """
    settings = get_settings()
    task_id = str(uuid.uuid4())[:8]

    # Resolve contest info for the download filename
    from app.models.contest import Contest
    from sqlalchemy import select as sa_select
    ct = await db.execute(sa_select(Contest).where(Contest.id == contest_id))
    contest = ct.scalar_one_or_none()
    contest_title = contest.title.replace('/', '_').replace('\\', '_') if contest else f"赛事{contest_id}"
    type_label = "报名数据" if export_type == "registration" else "成绩数据"
    download_filename = f"{contest_title}_{type_label}.xlsx"

    _export_tasks[task_id] = {
        "status": "processing", "file_path": None,
        "created_at": datetime.now(timezone.utc), "filename": download_filename,
    }

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = type_label

        # Resolve default field names after we have the contest
        if not field_names:
            score_cats = contest.score_categories if contest else []
            field_names = _resolve_default_field_names(export_type, score_cats)

        _write_header_row(ws, field_names)

        if export_type == "registration":
            rows = await get_registrations_for_export(db, contest_id, group_ids)
            _write_registration_rows(ws, rows, field_names)
        else:
            items, _ = await list_results(db, ResultFilter(contest_id=contest_id, page=1, page_size=settings.export_max_rows))
            await _write_result_rows(ws, items, db, field_names)

        _auto_fit_columns(ws)

        os.makedirs(settings.export_dir, exist_ok=True)
        filename = f"export_{task_id}_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(settings.export_dir, filename)
        wb.save(file_path)

        _export_tasks[task_id]["status"] = "completed"
        _export_tasks[task_id]["file_path"] = file_path

    except Exception as e:
        _export_tasks[task_id] = {
            "status": "failed", "file_path": None,
            "error": str(e), "created_at": _export_tasks[task_id]["created_at"],
        }

    return task_id


def get_export_task_status(task_id: str) -> dict | None:
    """Return the current status and file path for an export task, or None if the task_id is unknown."""
    return _export_tasks.get(task_id)
